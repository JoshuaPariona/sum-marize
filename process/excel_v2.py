""" Process notes that come in EXCEL """


import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import List, Dict, Tuple


def parse_excel(file_path: str) -> List[Dict]:
    """
    Parsea el contenido del archivo Excel y convierte cada fila en un diccionario con sus columnas.
    Además, identifica dinámicamente la cabecera con openpyxl.

    :param file_path: Ruta al archivo Excel.
    :return: Lista de diccionarios con los datos procesados.
    """
    # Usar openpyxl para cargar el archivo y buscar cabeceras
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Identificar la cabecera
    header_row, header_col = identify_table_headers(sheet)

    # Eliminar filas y columnas vacías y ajustar la tabla
    cleaned_data = adjust_table(sheet, header_row, header_col)

    # Convertir los datos limpios a DataFrame de pandas
    # Primera fila como cabecera
    df_adjusted = pd.DataFrame(cleaned_data[1:], columns=cleaned_data[0])

    # Renombrar columnas duplicadas
    df_adjusted = df_adjusted.loc[:, ~df_adjusted.columns.duplicated()]

    # Convertir los datos a una lista de diccionarios (JSON), omitiendo las cabeceras
    notes = df_adjusted.dropna(how="all").to_dict(orient="records")

    # Limpiar y normalizar los datos
    cleaned_notes = [clean_data(row) for row in notes[1:]]

    return cleaned_notes


def clean_data(row: Dict) -> Dict:
    """
    Limpia y normaliza los datos del Excel, asegurando que los tipos de datos sean compatibles con JSON.
    """
    cleaned_row = {}
    for key, value in row.items():
        if pd.isnull(value):  # Si el valor es NaN, lo manejamos como una cadena vacía
            cleaned_row[key] = ""  # Cambiar a "" para NaN
        elif isinstance(value, (np.int64, np.float64)):
            print(value)
            cleaned_row[key] = value.item()  # Convertir a int o float nativo
        else:
            # Intentar convertir a un número si es posible
            try:
                cleaned_row[key] = int(value) if isinstance(value, (int, float, str)) and str(
                    value).replace('.', '', 1).isdigit() else str(value).strip()
            except (ValueError, TypeError):
                cleaned_row[key] = str(value).strip() if value else ""

    return cleaned_row


def identify_table_headers(sheet) -> Tuple[int, int]:
    """
    Identifica la posición de la cabecera de la tabla en el archivo Excel usando openpyxl.
    Busca la fila y columna donde se encuentran los encabezados.

    :param sheet: Hoja activa del archivo Excel.
    :return: Una tupla (i, j) donde i es la fila y j es la columna en la que comienza la cabecera.
    """
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        for j, cell in enumerate(row):
            # Detectar una celda no vacía que parezca una cabecera
            if isinstance(cell, str) and cell.strip():
                return i, j  # Encontramos la cabecera en la posición (i, j)

    # Si no se encuentra una cabecera, se retorna por defecto (0, 0)
    return 0, 0


def adjust_table(sheet, header_row: int, header_col: int) -> List[List]:
    """
    Ajusta la tabla eliminando filas y columnas vacías antes de la cabecera, usando openpyxl.

    :param sheet: Hoja activa del archivo Excel.
    :param header_row: Fila en la que comienza la cabecera.
    :param header_col: Columna en la que comienza la cabecera.
    :return: Lista bidimensional con los datos limpios.
    """
    data = []

    # Recorrer las filas desde la fila de la cabecera
    for i, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True)):
        # Limpiar filas vacías
        # Ignorar las columnas antes de la cabecera
        row_data = row[header_col:]
        if any(cell is not None for cell in row_data):  # Si la fila no está completamente vacía
            data.append(row_data)

    # Extraer la fila de la cabecera y agregarla al principio
    header = [cell for cell in sheet.iter_rows(
        min_row=header_row + 1, max_row=header_row + 1, values_only=True)][0][header_col:]
    data.insert(0, header)

    return data
