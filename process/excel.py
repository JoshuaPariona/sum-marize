""" Process notes that come in EXCEL """

import openai
import os
from dotenv import load_dotenv
from typing import List, Dict
from openpyxl import load_workbook

load_dotenv()

SECRET_KEY = os.getenv('KEY_API_OPENAI')

# Clave de OpenAI
_api_key = SECRET_KEY

openai.api_key = _api_key

keywords = {"Código Alumno", "Apellidos y Nombres", "NP", "EV", "NF"}


def call_openai_for_mapping(headers: List[str]) -> Dict[str, str]:
    """
    Llama a la API de OpenAI para mapear encabezados a las palabras clave más relevantes usando la nueva API.
    """
    prompt = f"""
    Tengo los siguientes encabezados de un archivo Excel: {headers}. 
    Relaciónalos con las siguientes palabras clave: ['Código Alumno', 'Apellidos y Nombres', 'NP', 'EV', 'NF']. 
    Si algún encabezado no coincide directamente, intenta sugerir cuál palabra clave podría corresponder o indícame "Sin Coincidencia".
    """

    mensaje = [
        {"role": "system", "content": "Eres un asistente que mapea encabezados de Excel."},
        {"role": "user", "content": prompt}
    ]

    try:
        # Usar la nueva API ChatCompletion
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",  # O usa gpt-4 si tienes acceso
            messages=mensaje,
            max_tokens=150,
            temperature=0.5
        )

        # Procesar la respuesta
        mapping_suggestion = response['choices'][0]['message']['content'].strip(
        )

        # Intentar convertir el texto en un diccionario (ajusta según el formato de respuesta)
        try:
            mapping_dict = eval(mapping_suggestion)
        except:
            mapping_dict = {}  # Manejo de errores si eval no puede parsear la respuesta correctamente

        return mapping_dict

    except Exception as e:
        print(f"Error llamando a OpenAI: {e}")
        return {}


def parse_excel(file_path: str) -> List[Dict]:
    """
    Parsea el contenido del archivo Excel y convierte cada fila en un diccionario con sus columnas.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Obtener los encabezados de la primera fila del Excel
    headers = [cell.value for cell in sheet[1]]

    # Llama a OpenAI para mapear los encabezados a las palabras clave
    header_mapping = call_openai_for_mapping(headers)
    notes = []

    # Procesa las filas siguientes (omitiendo la primera que son los encabezados)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        cleaned_row = clean_data(row_dict, header_mapping)
        notes.append(cleaned_row)

    return notes


def clean_data(row: Dict, mapping: Dict) -> Dict:
    """
    Limpia y normaliza los datos del Excel, y los mapea a las palabras clave utilizando OpenAI.
    """
    cleaned_row = {mapping.get(key, key): str(
        value).strip() if value else "" for key, value in row.items()}
    return cleaned_row
